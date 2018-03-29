import openpyxl
import re

print("Введите название файла со список участников, без расширения файла:", )
# Name_Workbook_Competitor = input() + ".xlsx"
Name_Workbook_Competitor = 'First_step_06_05_2017_first_year.xlsx'

print(Name_Workbook_Competitor)

Workbook_Competitor = openpyxl.load_workbook(Name_Workbook_Competitor)
Worksheet_Competitor = Workbook_Competitor['Участники']

# Выбор участников соревнований по видам соревнований
Person = {}

Qty_String = 2
while Worksheet_Competitor.cell(row=Qty_String, column=2).value != None:

    NamePerson = Worksheet_Competitor.cell(row=Qty_String, column=2).value
    SexPerson = Worksheet_Competitor.cell(row=Qty_String, column=7).value
    GroupPerson = Worksheet_Competitor.cell(row=Qty_String, column=11).value
    AgePerson = Worksheet_Competitor.cell(row=Qty_String, column=8).value
    WeightPerson = Worksheet_Competitor.cell(row=Qty_String, column=10).value
    DzunroPerson = Worksheet_Competitor.cell(row=Qty_String, column=9).value
    KataPerson = Worksheet_Competitor.cell(row=Qty_String, column=11).value
    KumitePerson = Worksheet_Competitor.cell(row=Qty_String, column=12).value
    TeamPerson = Worksheet_Competitor.cell(row=Qty_String, column=1).value
    BirthdayPerson = Worksheet_Competitor.cell(row=Qty_String, column=3).value
    KuDanPerson = Worksheet_Competitor.cell(row=Qty_String, column=4).value
    SportCategoryPerson = Worksheet_Competitor.cell(row=Qty_String, column=5).value
    CoachPerson = Worksheet_Competitor.cell(row=Qty_String, column=6).value

    Person[Qty_String - 2] = {'id': Qty_String - 2,
                              'DataPerson': {'name': NamePerson, 'sex': SexPerson, 'group': GroupPerson,
                                             'age': AgePerson, 'weight': WeightPerson, 'dzunro': DzunroPerson,
                                             'kata': KataPerson, 'kumite': KumitePerson, 'team': TeamPerson,
                                             'birthday': BirthdayPerson, 'KuDan': KuDanPerson,'sportCategory': SportCategoryPerson,
                                             'coach': CoachPerson}}

    Qty_String += 1

# print(Person.items(), end='\n')
print("Всего участников: " + str(len(Person)))
print("====================================")

Female_A_4_Kata = {}
Female_B_4_Kata = {}
Female_A_4_Kumite = {}
Female_B_4_Kumite = {}

Female_A_5_Kata = {}
Female_B_5_Kata = {}
Female_A_5_Kumite = {}
Female_B_5_Kumite = {}

Female_A_6_Kata = {}
Female_B_6_Kata = {}
Female_A_6_Kumite = {}
Female_B_6_Kumite = {}

Female_A_7_Kata = {}
Female_B_7_Kata = {}
Female_A_7_Kumite = {}
Female_B_7_Kumite = {}

Female_A_89_Kata = {}
Female_B_89_Kata = {}
Female_A_89_Kumite = {}
Female_B_89_Kumite = {}

Female_A_1011_Kata = {}
Female_B_1011_Kata = {}
Female_A_1011_Kumite = {}
Female_B_1011_Kumite = {}

Female_A_1213_Kata = {}
Female_B_1213_Kata = {}
Female_A_1213_Kumite = {}
Female_B_1213_Kumite = {}

Female_A_1415_Kata = {}
Female_B_1415_Kata = {}
Female_A_1415_Kumite = {}
Female_B_1415_Kumite = {}

Male_A_4_Kata = {}
Male_B_4_Kata = {}
Male_A_4_Kumite = {}
Male_B_4_Kumite = {}

Male_A_5_Kata = {}
Male_B_5_Kata = {}
Male_A_5_Kumite = {}
Male_B_5_Kumite = {}

Male_A_6_Kata = {}
Male_B_6_Kata = {}
Male_A_6_Kumite_00 = {}
Male_B_6_Kumite_00 = {}
Male_A_6_Kumite_25 = {}
Male_B_6_Kumite_25 = {}

Male_A_7_Kata = {}
Male_B_7_Kata = {}
Male_A_7_Kumite_00 = {}
Male_B_7_Kumite_00 = {}
Male_A_7_Kumite_25 = {}
Male_B_7_Kumite_25 = {}

Male_A_89_Kata = {}
Male_B_89_Kata = {}
Male_A_89_Kumite_00 = {}
Male_B_89_Kumite_00 = {}
Male_A_89_Kumite_30 = {}
Male_B_89_Kumite_30 = {}

Male_A_1011_Kata = {}
Male_B_1011_Kata = {}
Male_A_1011_Kumite_00 = {}
Male_B_1011_Kumite_00 = {}
Male_A_1011_Kumite_35 = {}
Male_B_1011_Kumite_35 = {}

Male_A_1213_Kata = {}
Male_B_1213_Kata = {}
Male_A_1213_Kumite_00 = {}
Male_B_1213_Kumite_00 = {}
Male_A_1213_Kumite_45 = {}
Male_B_1213_Kumite_45 = {}

Male_A_1415_Kata = {}
Male_B_1415_Kata = {}
Male_A_1415_Kumite_00 = {}
Male_B_1415_Kumite_00 = {}
Male_A_1415_Kumite_55 = {}
Male_B_1415_Kumite_55 = {}

MF_A_911_Dzunro = {}
MF_A_1215_Dzunro = {}

listGroups = [Female_A_4_Kata, Female_B_4_Kata, Female_A_4_Kumite, Female_B_4_Kumite,
              Female_A_5_Kata, Female_B_5_Kata, Female_A_5_Kumite, Female_B_5_Kumite,
              Female_A_6_Kata, Female_B_6_Kata, Female_A_6_Kumite, Female_B_6_Kumite,
              Female_A_7_Kata, Female_B_7_Kata, Female_A_7_Kumite, Female_B_7_Kumite,
              Female_A_89_Kata, Female_B_89_Kata, Female_A_89_Kumite, Female_B_89_Kumite,
              Female_A_1011_Kata, Female_B_1011_Kata, Female_A_1011_Kumite, Female_B_1011_Kumite,
              Female_A_1213_Kata, Female_B_1213_Kata, Female_A_1213_Kumite, Female_B_1213_Kumite,
              Female_A_1415_Kata, Female_B_1415_Kata, Female_A_1415_Kumite,Female_B_1415_Kumite,
              Male_A_4_Kata, Male_B_4_Kata, Male_A_4_Kumite, Male_B_4_Kumite,
              Male_A_5_Kata, Male_B_5_Kata, Male_A_5_Kumite, Male_B_5_Kumite,
              Male_A_6_Kata, Male_B_6_Kata, Male_A_6_Kumite_00, Male_B_6_Kumite_00, Male_A_6_Kumite_25, Male_B_6_Kumite_25,
              Male_A_7_Kata, Male_B_7_Kata, Male_A_7_Kumite_00, Male_B_7_Kumite_00, Male_A_7_Kumite_25, Male_B_7_Kumite_25,
              Male_A_89_Kata, Male_B_89_Kata, Male_A_89_Kumite_00, Male_B_89_Kumite_00, Male_A_89_Kumite_30, Male_B_89_Kumite_30,
              Male_A_1011_Kata, Male_B_1011_Kata, Male_A_1011_Kumite_00, Male_B_1011_Kumite_00, Male_A_1011_Kumite_35, Male_B_1011_Kumite_35,
              Male_A_1213_Kata, Male_B_1213_Kata, Male_A_1213_Kumite_00, Male_B_1213_Kumite_00, Male_A_1213_Kumite_45, Male_B_1213_Kumite_45,
              Male_A_1415_Kata, Male_B_1415_Kata, Male_A_1415_Kumite_00, Male_B_1415_Kumite_00, Male_A_1415_Kumite_55, Male_B_1415_Kumite_55,
              MF_A_911_Dzunro, MF_A_1215_Dzunro]

for i in range(len(Person)):

    Name = Person[i]['DataPerson']['name']
    Sex = Person[i]['DataPerson']['sex']
    Group = Person[i]['DataPerson']['group']
    Age = Person[i]['DataPerson']['age']
    Weight = Person[i]['DataPerson']['weight']
    Dzunro = Person[i]['DataPerson']['dzunro']
    Kata = Person[i]['DataPerson']['kata']
    Kumite = Person[i]['DataPerson']['kumite']

    # Девочки 4 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 4:
        Female_A_4_Kata[i] = Person[i]
    if Sex == "ж" and (Kata == "Б") and Age == 4:
        Female_B_4_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 4:
        Female_A_4_Kumite[i] = Person[i]
    if Sex == "ж" and (Kumite == "Б") and Age == 4:
        Female_B_4_Kumite[i] = Person[i]

    # Девочки 5 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 5:
        Female_A_5_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 5:
        Female_B_5_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 5:
        Female_A_5_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 5:
        Female_B_5_Kumite[i] = Person[i]

    # Девочки 6 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 6:
        Female_A_6_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 6:
        Female_B_6_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 6:
        Female_A_6_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 6:
        Female_B_6_Kumite[i] = Person[i]

    # Девочки 7 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 7:
        Female_A_7_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 7:
        Female_B_7_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 7:
        Female_A_7_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 7:
        Female_B_7_Kumite[i] = Person[i]

    # Девочки 8-9 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 8 or Age == 9):
        Female_A_89_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 8 or Age == 9):
        Female_B_89_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 8 or Age == 9):
        Female_A_89_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 8 or Age == 9):
        Female_B_89_Kumite[i] = Person[i]

    # Девочки 10-11 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 10 or Age == 11):
        Female_A_1011_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 10 or Age == 11):
        Female_B_1011_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 10 or Age == 11):
        Female_A_1011_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 10 or Age == 11):
        Female_B_1011_Kumite[i] = Person[i]

    # Девочки 12-13 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 12 or Age == 13):
        Female_A_1213_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 12 or Age == 13):
        Female_B_1213_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 12 or Age == 13):
        Female_A_1213_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 12 or Age == 13):
        Female_B_1213_Kumite[i] = Person[i]

    # Девушки 14-15 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 14 or Age == 15):
        Female_A_1415_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 14 or Age == 15):
        Female_B_1415_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 14 or Age == 15):
        Female_A_1415_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 14 or Age == 15):
        Female_B_1415_Kumite[i] = Person[i]

    # Мальчики 4 года
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 4:
        Male_A_4_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 4:
        Male_B_4_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 4:
        Male_A_4_Kumite[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 4:
        Male_B_4_Kumite[i] = Person[i]

    # Мальчики 5 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 5:
        Male_A_5_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 5:
        Male_B_5_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 5:
        Male_A_5_Kumite[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 5:
        Male_B_5_Kumite[i] = Person[i]

    # Мальчики 6 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 6:
        Male_A_6_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 6:
        Male_B_6_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 6 and Weight <= 25:
        Male_A_6_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 6 and Weight <= 25:
        Male_B_6_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 6 and Weight > 25:
        Male_A_6_Kumite_25[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 6 and Weight > 25:
        Male_B_6_Kumite_25[i] = Person[i]

    # Мальчики 7 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 7:
        Male_A_7_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 7:
        Male_B_7_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 7 and Weight <= 25:
        Male_A_7_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 7 and Weight <= 25:
        Male_B_7_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 7 and Weight > 25:
        Male_A_7_Kumite_25[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 7 and Weight > 25:
        Male_B_7_Kumite_25[i] = Person[i]

    # Мальчики 8-9 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and (Age == 8 or Age == 9):
        Male_A_89_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 8 or Age == 9):
        Male_B_89_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 8 and Weight <= 30:
        Male_A_89_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 8 or Age == 9) and Weight <= 30:
        Male_B_89_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and (Age == 8 or Age == 9) and Weight > 30:
        Male_A_89_Kumite_30[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 8 or Age == 9) and Weight > 30:
        Male_B_89_Kumite_30[i] = Person[i]

    # Мальчики 10-11 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age == 10 or Age == 11):
        Male_A_1011_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 10 or Age == 11):
        Male_B_1011_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 10 or Age == 11) and Weight <= 35:
        Male_A_1011_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 10 or Age == 11) and Weight <= 35:
        Male_B_1011_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 10 or Age == 11) and Weight > 35:
        Male_A_1011_Kumite_35[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 10 or Age == 11) and Weight > 35:
        Male_B_1011_Kumite_35[i] = Person[i]

    # Мальчики 12-13 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age == 12 or Age == 13):
        Male_A_1213_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 12 or Age == 13):
        Male_B_1213_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 12 or Age == 13) and Weight <= 45:
        Male_A_1213_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 12 or Age == 13) and Weight <= 45:
        Male_B_1213_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 12 or Age == 13) and Weight > 45:
        Male_A_1213_Kumite_45[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 12 or Age == 13) and Weight > 45:
        Male_A_1213_Kumite_45[i] = Person[i]

    # Юноши 14-15 лет
    if Sex == "ч"  and (Kata == "A" or Kata == "А") and (Age == 14 or Age == 15):
        Male_A_1415_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 14 or Age == 15):
        Male_B_1415_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 14 or Age == 15) and Weight <= 55:
        Male_A_1415_Kumite_00[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 14 or Age == 15) and Weight <= 55:
        Male_B_1415_Kumite_00[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 14 or Age == 15) and Weight > 55:
        Male_A_1415_Kumite_55[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 14 or Age == 15) and Weight > 55:
        Male_B_1415_Kumite_55[i] = Person[i]

   # Джунро мужчины
    if Age >= 9 and Age <= 11 and Dzunro == "Д":
        MF_A_911_Dzunro[i] = Person[i]
    if Age >= 12 and Age <= 15 and Dzunro == "Д":
        MF_A_1215_Dzunro[i] = Person[i]

# функция печати кол-ва участников указанной группы
def participant_print(participant_list_str,
                      participant_list):  # передаем название списка участников в виде строки и ссылку на сам словарь с участниками
    data = re.split('_', participant_list_str)  # ['Femail', 'A', '6', 'Kata']
    if data[0] == 'MF':
        data[0] = 'Котен ката'
        print(data[0] + ", " + data[2] + " лет")
        print("Кол-во участников: " + str(len(participant_list)))
        print(participant_list)
        return
    elif data[0] == 'Female':
        data[0] = 'Женская'
    elif data[0] == 'Male':
        data[0] = 'Мужская'
    print(data[0] + " группа " + data[1] + " " + data[2] + " " + data[3])
    print("Кол-во участников: " + str(len(participant_list)))
    print(participant_list)
    print("====================================")

# передаем название списка участников в виде строки и ссылку на сам словарь с участниками
# participant_print('Dzunro_911', Dzunro_911)


# функция создания excel файла для каждой группы участников
# передаем список участников и название файла, куда сохранить результаты
def create_olimp_list(list_competitors, save_name):
    global sheetNet
    wb = openpyxl.load_workbook(Name_Workbook_Competitor)
    sheet = wb['Вставка']

    if len(list_competitors) == 0:
        print("Участники в этой категории отсутствуют")
    else:
        print("Всего " + str(len(list_competitors)) + " участников в группе " + save_name + ".")
        data = re.split('_', save_name)  # ['Femail', 'A', '6', 'Kata', 'Weight']

# выбор нужной сетки
        if len(list_competitors) > 0 and len(list_competitors) < 5:
            sheetNet = wb['4']
        elif len(list_competitors) > 4 and len(list_competitors) < 9:
            sheetNet = wb['8']
        elif len(list_competitors) > 8 and len(list_competitors) < 17:
            sheetNet = wb['16']
        elif len(list_competitors) > 16 and len(list_competitors) < 33:
            sheetNet = wb['32']
        elif len(list_competitors) > 32 and len(list_competitors) < 64:
            sheetNet = wb['64']
        elif len(list_competitors) > 64 and len(list_competitors) < 128:
            sheetNet = wb['128']
        # заполнение шаблона сеток
        if data[0] == 'Female' and int(data[2]) < int(12):
            data[0] = 'Девочки'
        elif data[0] == 'Female' and int(data[2]) > int(11) and int(data[2]) < int(18):
            data[0] = 'Девушки'
        elif data[0] == 'Female' and int(data[2]) > int(17):
            data[0] = 'Женщины'
        elif data[0] == 'Male' and int(data[2]) < int(12):
            data[0] = 'Мальчики'
        elif data[0] == 'Male' and int(data[2]) > int(11) and int(data[2]) < int(18):
            data[0] = 'Юноши'
        elif data[0] == 'Male' and int(data[2]) > int(17):
            data[0] = 'Мужчины'

        if data[3] == 'Kata':
            data[3] = 'Ката_1_year'
        elif data[3] == 'Kumite':
            data[3] = 'Кумитэ_1_year'
        if data[3] == 'Dzunro':
            data[3] = 'Джунро'

#       if data[3] == 'Кумитэ' and Person[key]['DataPerson']['weight']:
#           data[4] = 'легкие'
#       else:
#           data[4] = 'тяжелые'

        sheetNet['C2'] = data[0]
        sheetNet['E2'] = data[1]
        sheetNet['G2'] = data[3]
        sheetNet['I2'] = data[2]
#       sheetNet['K2'] = data[4]

        j = 2
        for key in list_competitors:
            # print (str(key))
            cell = 'C' + str(j)
            print(Person[key]['DataPerson']['name'])
            sheet[cell] = Person[key]['DataPerson']['name']
            cell = 'B' + str(j)
            sheet[cell] = Person[key]['DataPerson']['team']
            cell = 'D' + str(j)
            sheet[cell] = Person[key]['DataPerson']['birthday']
            cell = 'E' + str(j)
            sheet[cell] = Person[key]['DataPerson']['KuDan']
            cell = 'F' + str(j)
            sheet[cell] = Person[key]['DataPerson']['sportCategory']
            cell = 'G' + str(j)
            sheet[cell] = Person[key]['DataPerson']['coach']
            j += 1

        wb.save(save_name + '_first_year'+'.xlsx')
 #   print("Готово !")
    print("_____________________________________________")

# передаем список участников и название файла, куда сохранить результаты
# create_olimp_list(Male_A_1415_Kata, 'test')


# надо завернуть кусок кода ниже в отдельную функцию
# def create_all_olimp_list():
a = 1
for j in range(len(listGroups)):
    if len(listGroups[j]) > 0:
         # print(str(a)+") " + next((k for k, v in locals().items() if id(listGroups[j]) == id(v))))
         create_olimp_list(listGroups[j], next((k for k, v in locals().items() if id(listGroups[j]) == id(v))))
         a += 1
#     print("опачки")

# create_all_olimp_list()
