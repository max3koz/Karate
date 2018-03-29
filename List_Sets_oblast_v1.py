import openpyxl
import re

print("Введите название файла со список участников, без расширения файла:", )
Name_Workbook_Competitor = input() + ".xlsx"
#Name_Workbook_Competitor = 'Сетки на 28.10.2017.xlsx'

print(Name_Workbook_Competitor)

Workbook_Competitor = openpyxl.load_workbook(Name_Workbook_Competitor)
Worksheet_Competitor = Workbook_Competitor['Участники']

# Выбор участников соревнований по видам соревнований
Person = {}

Qty_String = 2
while Worksheet_Competitor.cell(row=Qty_String, column=2).value != None:

    NamePerson = Worksheet_Competitor.cell(row=Qty_String, column=2).value
    SexPerson = Worksheet_Competitor.cell(row=Qty_String, column=7).value
    GroupPerson = Worksheet_Competitor.cell(row=Qty_String, column=11).value
    AgePerson = Worksheet_Competitor.cell(row=Qty_String, column=8).value
    WeightPerson = Worksheet_Competitor.cell(row=Qty_String, column=10).value
    DzunroPerson = Worksheet_Competitor.cell(row=Qty_String, column=9).value
    KataPerson = Worksheet_Competitor.cell(row=Qty_String, column=11).value
    KumitePerson = Worksheet_Competitor.cell(row=Qty_String, column=12).value
    TeamPerson = Worksheet_Competitor.cell(row=Qty_String, column=1).value
    BirthdayPerson = Worksheet_Competitor.cell(row=Qty_String, column=3).value
    KuDanPerson = Worksheet_Competitor.cell(row=Qty_String, column=4).value
    SportCategoryPerson = Worksheet_Competitor.cell(row=Qty_String, column=5).value
    CoachPerson = Worksheet_Competitor.cell(row=Qty_String, column=6).value

    Person[Qty_String - 2] = {'id': Qty_String - 2,
                              'DataPerson': {'name': NamePerson, 'sex': SexPerson, 'group': GroupPerson,
                                             'age': AgePerson, 'weight': WeightPerson, 'dzunro': DzunroPerson,
                                             'kata': KataPerson, 'kumite': KumitePerson, 'team': TeamPerson,
                                             'birthday': BirthdayPerson, 'KuDan': KuDanPerson,'sportCategory': SportCategoryPerson,
                                             'coach': CoachPerson}}

    Qty_String += 1

# print(Person.items(), end='\n')
print("Всего участников: " + str(len(Person)))
print("====================================")

Female_B_5_Kata = {}
Female_B_5_Kumite = {}

Female_A_6_Kata = {}
Female_B_6_Kata = {}
Female_A_6_Kumite = {}
Female_B_6_Kumite = {}

Female_A_7_Kata = {}
Female_B_7_Kata = {}
Female_A_7_Kumite = {}
Female_B_7_Kumite = {}

Female_A_8_Kata = {}
Female_B_8_Kata = {}
Female_A_8_Kumite = {}
Female_B_8_Kumite = {}

Female_A_9_Kata = {}
Female_B_9_Kata = {}
Female_A_9_Kumite = {}
Female_B_9_Kumite = {}

Female_A_1011_Kata = {}
Female_B_1011_Kata = {}
Female_A_1011_Kumite = {}
Female_B_1011_Kumite = {}

Female_A_1213_Kata = {}
Female_B_1213_Kata = {}
Female_A_1213_Kumite = {}
Female_B_1213_Kumite = {}

Female_A_1415_Kata = {}
Female_B_1415_Kata = {}
Female_A_1415_Kumite = {}
Female_B_1415_Kumite = {}

Female_A_1617_Kata = {}
Female_B_1617_Kata = {}
Female_A_1617_Kumite = {}
Female_B_1617_Kumite = {}

Female_A_18_Kata = {}
Female_B_18_Kata = {}
Female_A_18_Kumite = {}
Female_B_18_Kumite = {}

Female_A_35_Kata = {}


Male_B_5_Kata = {}
Male_B_5_Kumite_0 = {}

Male_A_6_Kata = {}
Male_B_6_Kata = {}
Male_A_6_Kumite_0 = {}
Male_B_6_Kumite_0 = {}
Male_A_6_Kumite_22 = {}
Male_B_6_Kumite_22 = {}

Male_A_7_Kata = {}
Male_B_7_Kata = {}
Male_A_7_Kumite_0 = {}
Male_B_7_Kumite_0 = {}
Male_A_7_Kumite_25 = {}
Male_B_7_Kumite_25 = {}

Male_A_8_Kata = {}
Male_B_8_Kata = {}
Male_A_8_Kumite_0 = {}
Male_B_8_Kumite_0 = {}
Male_A_8_Kumite_28 = {}
Male_B_8_Kumite_28 = {}

Male_A_9_Kata = {}
Male_B_9_Kata = {}
Male_A_9_Kumite_0 = {}
Male_B_9_Kumite_0 = {}
Male_A_9_Kumite_30 = {}
Male_B_9_Kumite_30 = {}

Male_A_1011_Kata = {}
Male_B_1011_Kata = {}
Male_A_1011_Kumite_0 = {}
Male_B_1011_Kumite_0 = {}
Male_A_1011_Kumite_37 = {}
Male_B_1011_Kumite_37 = {}

Male_A_1213_Kata = {}
Male_B_1213_Kata = {}
Male_A_1213_Kumite_0 = {}
Male_B_1213_Kumite_0 = {}
Male_A_1213_Kumite_46 = {}
Male_B_1213_Kumite_46 = {}

Male_A_1415_Kata = {}
Male_B_1415_Kata = {}
Male_A_1415_Kumite_0 = {}
Male_B_1415_Kumite_0 = {}
Male_A_1415_Kumite_60 = {}

Male_A_1617_Kata = {}
Male_B_1617_Kata = {}
Male_A_1617_Kumite_0 = {}
Male_B_1617_Kumite_0 = {}
Male_A_1617_Kumite_68 = {}

Male_A_18_Kata = {}
Male_B_18_Kata = {}
Male_A_18_Kumite_0 = {}
Male_B_18_Kumite_0 = {}

Male_A_35_Kata = {}
Male_A_35_Kumite_0 = {}

Male_A_911_Koten = {}
Male_A_1215_Koten = {}
Male_A_16_Koten = {}

listGroups = [Female_B_5_Kata, Female_B_5_Kumite, Female_A_6_Kata,Female_B_6_Kata, Female_A_6_Kumite,
              Female_B_6_Kumite, Female_A_7_Kata, Female_B_7_Kata, Female_A_7_Kumite, Female_B_7_Kumite,
              Female_A_8_Kata, Female_B_8_Kata, Female_A_8_Kumite, Female_B_8_Kumite, Female_A_9_Kata,
              Female_B_9_Kata, Female_A_9_Kumite, Female_B_9_Kumite, Female_A_1011_Kata, Female_B_1011_Kata,
              Female_A_1011_Kumite, Female_B_1011_Kumite, Female_A_1213_Kata, Female_B_1213_Kata, Female_A_1213_Kumite,
              Female_B_1213_Kumite, Female_A_1415_Kata, Female_B_1415_Kata, Female_A_1415_Kumite, Female_B_1415_Kumite,
              Female_A_1617_Kata, Female_B_1617_Kata, Female_A_1617_Kumite, Female_B_1617_Kumite, Female_A_18_Kata,
              Female_B_18_Kata, Female_A_18_Kumite, Female_B_18_Kumite, Female_A_35_Kata, Male_B_5_Kata, Male_B_5_Kumite_0,
              Male_A_6_Kata, Male_B_6_Kata, Male_A_6_Kumite_0, Male_B_6_Kumite_0, Male_A_6_Kumite_22, Male_B_6_Kumite_22,
              Male_A_7_Kata, Male_B_7_Kata, Male_A_7_Kumite_0, Male_B_7_Kumite_0, Male_A_7_Kumite_25, Male_B_7_Kumite_25,
              Male_A_8_Kata, Male_B_8_Kata, Male_A_8_Kumite_0, Male_B_8_Kumite_0, Male_A_8_Kumite_28, Male_B_8_Kumite_28,
              Male_A_9_Kata, Male_B_9_Kata, Male_A_9_Kumite_0, Male_B_9_Kumite_0, Male_A_9_Kumite_30, Male_B_9_Kumite_30,
              Male_A_1011_Kata, Male_B_1011_Kata, Male_A_1011_Kumite_0, Male_B_1011_Kumite_0, Male_A_1011_Kumite_37,
              Male_B_1011_Kumite_37, Male_A_1213_Kata, Male_B_1213_Kata, Male_A_1213_Kumite_0, Male_B_1213_Kumite_0,
              Male_A_1213_Kumite_46, Male_B_1213_Kumite_46, Male_A_1415_Kata, Male_B_1415_Kata, Male_A_1415_Kumite_0,
              Male_B_1415_Kumite_0, Male_A_1415_Kumite_60, Male_A_1617_Kata, Male_B_1617_Kata, Male_A_1617_Kumite_0,
              Male_B_1617_Kumite_0, Male_A_1617_Kumite_68, Male_A_18_Kata, Male_B_18_Kata, Male_A_18_Kumite_0,
              Male_B_18_Kumite_0, Male_A_35_Kata, Male_A_35_Kumite_0, Male_A_911_Koten, Male_A_1215_Koten, Male_A_16_Koten]

for i in range(len(Person)):

    Name = Person[i]['DataPerson']['name']
    Sex = Person[i]['DataPerson']['sex']
    Group = Person[i]['DataPerson']['group']
    Age = Person[i]['DataPerson']['age']
    Weight = Person[i]['DataPerson']['weight']
    Dzunro = Person[i]['DataPerson']['dzunro']
    Kata = Person[i]['DataPerson']['kata']
    Kumite = Person[i]['DataPerson']['kumite']

    # Девочки до 6 лет
    if Sex == "ж" and (Kata == "Б" or Kata == "B") and Age < 6:
        Female_B_5_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "Б" or Kumite == "B") and Age < 6:
        Female_B_5_Kumite[i] = Person[i]

    # Девочки 6 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 6:
        Female_A_6_Kata[i] = Person[i]
    if Sex == "ж" and (Kata == "Б" or Kata == "B") and Age == 6:
        Female_B_6_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 6:
        Female_A_6_Kumite[i] = Person[i]
    if Sex == "ж" and (Kumite == "Б" or Kumite == "B") and Age == 6:
        Female_B_6_Kumite[i] = Person[i]

    # Девочки 7 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 7:
        Female_A_7_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 7:
        Female_B_7_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 7:
        Female_A_7_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 7:
        Female_B_7_Kumite[i] = Person[i]

    # Девочки 8 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 8:
        Female_A_8_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 8:
        Female_B_8_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 8:
        Female_A_8_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 8:
        Female_B_8_Kumite[i] = Person[i]

    # Девочки 9 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age == 9:
        Female_A_9_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and Age == 9:
        Female_B_9_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and Age == 9:
        Female_A_9_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and Age == 9:
        Female_B_9_Kumite[i] = Person[i]

    # Девочки 10-11 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 10 or Age == 11):
        Female_A_1011_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 10 or Age == 11):
        Female_B_1011_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 10 or Age == 11):
        Female_A_1011_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 10 or Age == 11):
        Female_B_1011_Kumite[i] = Person[i]

    # Девочки 12-13 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 12 or Age == 13):
        Female_A_1213_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 12 or Age == 13):
        Female_B_1213_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 12 or Age == 13):
        Female_A_1213_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 12 or Age == 13):
        Female_B_1213_Kumite[i] = Person[i]

    # Девушки 14-15 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 14 or Age == 15):
        Female_A_1415_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 14 or Age == 15):
        Female_B_1415_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age == 14 or Age == 15):
        Female_A_1415_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 14 or Age == 15):
        Female_B_1415_Kumite[i] = Person[i]

    # Девушки 16-17 лет
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age == 16 or Age == 17):
        Female_A_1617_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age == 16 or Age == 17):
        Female_B_1617_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А"or Kumite == "A") and (Age == 16 or Age == 17):
        Female_A_1617_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age == 16 or Age == 17):
        Female_A_1617_Kumite[i] = Person[i]

    # Женщины 18-34 лет ката "А", 18+ ката "Б" и кумитэ обе категории
    if Sex == "ж" and (Kata == "А" or Kata == "A") and (Age >= 18 and Age < 35):
        Female_A_18_Kata[i] = Person[i]
    if Sex == "ж" and Kata == "Б" and (Age >= 18):
        Female_B_18_Kata[i] = Person[i]
    if Sex == "ж" and (Kumite == "А" or Kumite == "A") and (Age >= 18 and Age < 35):
        Female_A_18_Kumite[i] = Person[i]
    if Sex == "ж" and Kumite == "Б" and (Age >= 18):
        Female_B_18_Kumite[i] = Person[i]

    # Женщины 35 лет ката "А"
    if Sex == "ж" and (Kata == "А" or Kata == "A") and Age >= 35:
        Female_A_35_Kata[i] = Person[i]

    # Мальчики до 6 лет
    if Sex == "ч" and (Kata == "Б" or Kata == "B") and Age < 6:
        Male_B_5_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "Б" or Kumite == "B") and Age < 6:
        Male_B_5_Kumite_0[i] = Person[i]

    # Мальчики 6 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 6:
        Male_A_6_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 6:
        Male_B_6_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 6 and Weight <= 22:
        Male_A_6_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 6 and Weight <= 22:
        Male_B_6_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 6 and Weight > 22:
        Male_A_6_Kumite_22[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 6 and Weight > 22:
        Male_B_6_Kumite_22[i] = Person[i]

    # Мальчики 7 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 7:
        Male_A_7_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 7:
        Male_B_7_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 7 and Weight <= 25:
        Male_A_7_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 7 and Weight <= 25:
        Male_B_7_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 7 and Weight > 25:
        Male_A_7_Kumite_25[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 7 and Weight > 25:
        Male_B_7_Kumite_25[i] = Person[i]

    # Мальчики 8 лет
    if Sex == "ч" and (Kata == "А" or Kata == "A") and Age == 8:
        Male_A_8_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 8:
        Male_B_8_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 8 and Weight <= 28:
        Male_A_8_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 8 and Weight <= 28:
        Male_B_8_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "А" or Kumite == "A") and Age == 8 and Weight > 28:
        Male_A_8_Kumite_28[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 8 and Weight > 28:
        Male_B_8_Kumite_28[i] = Person[i]

    # Мальчики 9 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and Age == 9:
        Male_A_9_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and Age == 9:
        Male_B_9_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and Age == 9 and Weight <= 30:
        Male_A_9_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 9 and Weight <= 30:
        Male_B_9_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and Age == 9 and Weight > 30:
        Male_A_9_Kumite_30[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and Age == 9 and Weight > 30:
        Male_B_9_Kumite_30[i] = Person[i]

    # Мальчики 10-11 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age == 10 or Age == 11):
        Male_A_1011_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 10 or Age == 11):
        Male_B_1011_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 10 or Age == 11) and Weight <= 37:
        Male_A_1011_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 10 or Age == 11) and Weight <= 37:
        Male_B_1011_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 10 or Age == 11) and Weight > 37:
        Male_A_1011_Kumite_37[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 10 or Age == 11) and Weight > 37:
        Male_B_1011_Kumite_37[i] = Person[i]

    # Мальчики 12-13 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age == 12 or Age == 13):
        Male_A_1213_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 12 or Age == 13):
        Male_B_1213_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 12 or Age == 13) and Weight <= 46:
        Male_A_1213_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 12 or Age == 13) and Weight <= 46:
        Male_B_1213_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 12 or Age == 13) and Weight > 46:
        Male_A_1213_Kumite_46[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 12 or Age == 13) and Weight > 46:
        Male_B_1213_Kumite_46[i] = Person[i]

    # Юноши 14-15 лет
    if Sex == "ч"  and (Kata == "A" or Kata == "А") and (Age == 14 or Age == 15):
        Male_A_1415_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 14 or Age == 15):
        Male_B_1415_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 14 or Age == 15) and Weight <= 60:
        Male_A_1415_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 14 or Age == 15):
        Male_B_1415_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 14 or Age == 15) and Weight > 60:
        Male_A_1415_Kumite_60[i] = Person[i]

    # Юноши 16-17 лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age == 16 or Age == 17):
        Male_A_1617_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age == 16 or Age == 17):
        Male_B_1617_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 16 or Age == 17) and Weight <= 68:
        Male_A_1617_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age == 16 or Age == 17):
        Male_B_1617_Kumite_0[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age == 16 or Age == 17) and Weight > 68:
        Male_A_1617_Kumite_68[i] = Person[i]

    # Мужчины 18-34 лет ката, кумитэ группа А, 18+ ката, кумитэ группа Б
    if Sex == "ч" and (Kata == "A" or Kata == "А") and (Age >= 18 and Age < 35):
        Male_A_18_Kata[i] = Person[i]
    if Sex == "ч" and Kata == "Б" and (Age >= 18):
        Male_B_18_Kata[i] = Person[i]
    if Sex == "ч" and (Kumite == "A" or Kumite == "А") and (Age >= 18 and Age < 35):
        Male_A_18_Kumite_0[i] = Person[i]
    if Sex == "ч" and Kumite == "Б" and (Age >= 18):
        Male_B_18_Kumite_0[i] = Person[i]

    # Мужчины 35+ лет
    if Sex == "ч" and (Kata == "A" or Kata == "А") and Age >= 35:
        Male_A_35_Kata[i] = Person[i]
    if Sex == "ч" and Kumite == "А" and Age >= 35:
        Male_A_35_Kumite_0[i] = Person[i]

    # Джунро мужчины
    if Age >= 4 and Age <= 11 and Dzunro == "А":
        Male_A_911_Koten[i] = Person[i]
    if Age >= 12 and Age <= 15 and Dzunro == "А":
        Male_A_1215_Koten[i] = Person[i]
    if Age >= 16 and Dzunro == "А":
        Male_A_16_Koten[i] = Person[i]


# функция печати кол-ва участников указанной группы
def participant_print(participant_list_str,
                      participant_list):  # передаем название списка участников в виде строки и ссылку на сам словарь с участниками
    data = re.split('_', participant_list_str)  # ['Femail', 'A', '6', 'Kata']
    if data[0] == 'Koten':
        data[0] = 'Котен'
        print(data[0] + " ката " + data[1])
        print("Кол-во участников: " + str(len(participant_list)))
        print(participant_list)
        return
    elif data[0] == 'Female':
        data[0] = 'Женская'
    elif data[0] == 'Male':
        data[0] = 'Мужская'
    print(data[0] + " группа " + data[1] + " " + data[2] + " " + data[3])
    print("Кол-во участников: " + str(len(participant_list)))
    print(participant_list)
    print("====================================")

# передаем название списка участников в виде строки и ссылку на сам словарь с участниками
# participant_print('Dzunro_911', Dzunro_911)


# функция создания excel файла для каждой группы участников
# передаем список участников и название файла, куда сохранить результаты
def create_olimp_list(list_competitors, save_name):
    global sheetNet
    wb = openpyxl.load_workbook(Name_Workbook_Competitor)
    sheet = wb['Вставка']

    if len(list_competitors) == 0:
        print("Участники в этой категории отсутствуют")
    else:
        print("Всего " + str(len(list_competitors)) + " участников в группе " + save_name + ".")
        data = re.split('_', save_name)  # ['Femail', 'A', '6', 'Kata', 'Weight']

# выбор нужной сетки
        if len(list_competitors) > 0 and len(list_competitors) < 5:
            sheetNet = wb['4']
        elif len(list_competitors) > 4 and len(list_competitors) < 9:
            sheetNet = wb['8']
        elif len(list_competitors) > 8 and len(list_competitors) < 17:
            sheetNet = wb['16']
        elif len(list_competitors) > 16 and len(list_competitors) < 33:
            sheetNet = wb['32']
        elif len(list_competitors) > 32 and len(list_competitors) < 64:
            sheetNet = wb['64']
        elif len(list_competitors) > 64 and len(list_competitors) < 128:
            sheetNet = wb['128']
        # заполнение шаблона сеток
        if data[0] == 'Female' and int(data[2]) < int(12):
            data[0] = 'Дівчата'
        elif data[0] == 'Female' and int(data[2]) > int(11) and int(data[2]) < int(18):
            data[0] = 'Дівчата'
        elif data[0] == 'Female' and int(data[2]) > int(17):
            data[0] = 'Жінки'
        elif data[0] == 'Male' and int(data[2]) < int(12):
            data[0] = 'Хлопці'
        elif data[0] == 'Male' and int(data[2]) > int(11) and int(data[2]) < int(18):
            data[0] = 'Юнаки'
        elif data[0] == 'Male' and int(data[2]) > int(17):
            data[0] = 'Чоловіки'

        if data[3] == 'Kata':
            data[3] = 'Ката'
        elif data[3] == 'Kumite':
            data[3] = 'Кумитэ'
        if data[3] == 'Koten':
            data[3] = 'Котен'

        if data[1] == "B":
            data[1] = 'Б'

#       if data[3] == 'Кумитэ' and Person[key]['DataPerson']['weight']:
#           data[4] = 'легкие'
#       else:
#           data[4] = 'тяжелые'

        sheetNet['C2'] = data[3]
        sheetNet['E2'] = data[1]
        sheetNet['G2'] = data[0]
        sheetNet['I2'] = data[2] + ' років. '
        if data[3] == 'Кумитэ' and (data[0] == 'Хлопці' or data[0] == 'Юнаки' or data[0] == 'Чоловіки'):
            sheetNet['K2'] = save_name
        #else: sheetNet['K2'] = 'Важкі'
        #sheetNet['K2'] = save_name

        j = 2
        for key in list_competitors:
            # print (str(key))
            cell = 'C' + str(j)
            print(Person[key]['DataPerson']['name'])
            sheet[cell] = Person[key]['DataPerson']['name']
            cell = 'B' + str(j)
            sheet[cell] = Person[key]['DataPerson']['team']
            cell = 'D' + str(j)
            sheet[cell] = Person[key]['DataPerson']['birthday']
            cell = 'E' + str(j)
            sheet[cell] = Person[key]['DataPerson']['KuDan']
            cell = 'F' + str(j)
            sheet[cell] = Person[key]['DataPerson']['sportCategory']
            cell = 'G' + str(j)
            sheet[cell] = Person[key]['DataPerson']['coach']
            j += 1

        wb.save(save_name + '.xlsx')
 #   print("Готово !")
    print("_____________________________________________")

# передаем список участников и название файла, куда сохранить результаты
# create_olimp_list(Male_A_1415_Kata, 'test')


# надо завернуть кусок кода ниже в отдельную функцию
# def create_all_olimp_list():
a = 1
for j in range(len(listGroups)):
    if len(listGroups[j]) > 0:
         # print(str(a)+") " + next((k for k, v in locals().items() if id(listGroups[j]) == id(v))))
         create_olimp_list(listGroups[j], next((k for k, v in locals().items() if id(listGroups[j]) == id(v))))
         a += 1
#     print("опачки")

# create_all_olimp_list()
